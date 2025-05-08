# main.py (Version 8.3 - Corrected Excel Export Syntax & Indentation)

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import json
import os
import platform
import webbrowser
from datetime import datetime
import re # Keep for potential future use

# --- Import necessary export libraries ---
try:
    import openpyxl
    from openpyxl.styles import Font as OpenpyxlFont, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.hyperlink import Hyperlink
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# --- Constants & Appearance ---
ctk.set_appearance_mode("Light") # Force Light mode for consistent background

# Define Brand Colors
PRIMARY_COLOR = "#39B54A" # Main Green
SECONDARY_COLOR = "#14467C" # Dark Blue
ACCENT_COLOR = "#5DC66A" # Lighter green for hover
TEXT_ON_PRIMARY = "#FFFFFF" # White text on green buttons/tabs
TEXT_ON_SECONDARY = "#FFFFFF" # White text on blue buttons
BACKGROUND_COLOR = "#FFFFFF" # Explicit white background
FRAME_BG_COLOR = "#FFFFFF" # Use white for frames too for cleaner look
DISABLED_COLOR = "#B0B0B0" # Color for disabled elements (used by CTk theme)
ERROR_COLOR = "#D32F2F" # Error text/indicator color
TEXT_COLOR_DARK = "#242424" # For default text
TEXT_COLOR_LIGHT = "#676767" # For secondary text (like status bar)

# Font Definitions using CTkFont
HEADER_FONT_FAMILY = "Arial Black"
BODY_FONT_FAMILY = "Segoe UI" # Common modern font
HEADER_FONT_SIZE = 22
BODY_FONT_SIZE_LARGE = 16
BODY_FONT_SIZE_MEDIUM = 13
BODY_FONT_SIZE_SMALL = 12
STATUS_FONT_SIZE = 11
LINK_ENTRY_FONT = ctk.CTkFont(family=BODY_FONT_FAMILY, size=BODY_FONT_SIZE_SMALL - 1) # Slightly smaller for link entry


# --- Checklist Structure (Unchanged) ---
CHECKLIST_STRUCTURE = [("Fire Safety Training", [("Have you commenced Fire Safety presentations as scheduled?", "yes_no", True), ("How are you tracking training completion?", "text", True)]), ("Documentation & Certifications", [("Is your Fire NOC valid and current?", "yes_no", True), ("Are warehouse fire layout diagrams displayed properly?", "yes_no", True)]), ("Safety Infrastructure", [("Have fluorescent markings been installed for emergency evacuation routes?", "yes_no", True), ("Are smoke detection systems, fire alarms, and emergency notification boards in place?", "yes_no", True), ("When was the last functionality test for sprinkler systems and fire hydrants?", "text", True)]), ("Operational Protocols", [("Is visitor registration being properly maintained?", "yes_no", True), ("Have daily SOPs and safety checklists been implemented?", "yes_no", True), ("How are you enforcing the prohibition of fire-ignition tools?", "text", True), ("Have you established machinery inspection schedules for hazard identification?", "yes_no", True)]), ("Maintenance Documentation", [("Has the procurement team implemented maintenance logbook protocols?", "yes_no", True)]), ("Personnel Qualification", [("Have you verified ITI certification or equivalent for all electrical personnel?", "yes_no", True)]), ("Safety Engagement Initiatives", [("What safety engagement activities have you organized recently?", "text", False), ("Which best practices from training have you implemented?", "text", False), ("Have you developed facility-specific internal safety protocols?", "yes_no", True)]), ("Compliance Verification", [("Have cross-Warehouse audits been conducted?", "yes_no", False), ("Is your monthly machinery safety inspection schedule established?", "yes_no", True), ("When was your last mock drill conducted?", "text", True), ("How are you maintaining inspection and compliance records?", "text", True)]), ("Seasonal Safety", [("Have all seasonal equipment (water coolers, etc.) been inspected?", "yes_no", True)])]


# ==============================================================================
# Main Application Class
# ==============================================================================
class WarehouseSafetyApp(ctk.CTk):
    """Main application window."""
    def __init__(self):
        super().__init__(fg_color=BACKGROUND_COLOR)
        self.title("Warehouse Safety Checklist Application")
        self.geometry("1100x850") # Back to slightly smaller width
        self.minsize(950, 750)

        # --- Data Storage Initialization ---
        self.project_file_path = None
        self.metadata_vars = {k: tk.StringVar() for k in ["Warehouse Name", "Location", "Report Date", "Report Month", "Uploaded By Name", "Uploaded By Role", "Uploaded By Emp ID", "Uploaded By Email", "Manager Name"]}
        self.metadata_vars["Report Date"].set(datetime.now().strftime('%Y-%m-%d'))
        self.metadata_vars["Report Month"].set(datetime.now().strftime('%B %Y'))
        # **REVERTED**: Simple structure for checklist answers only
        self.checklist_data_vars = {}
        self.near_miss_vars = {k: tk.StringVar() for k in ["Incident Date", "Incident Location", "Description", "Immediate Action", "Prevention Suggestion"]}
        self.near_miss_attachments = [] # URLs
        self.action_points_text_var = tk.StringVar()
        self.general_attachments = [] # URLs
        self.status_var = tk.StringVar()

        # --- Define CTkFonts ---
        self.header_font = ctk.CTkFont(family=HEADER_FONT_FAMILY, size=HEADER_FONT_SIZE, weight="bold")
        self.section_header_font = ctk.CTkFont(family=BODY_FONT_FAMILY, size=BODY_FONT_SIZE_LARGE, weight="bold")
        self.question_font = ctk.CTkFont(family=BODY_FONT_FAMILY, size=BODY_FONT_SIZE_MEDIUM, weight="bold")
        self.answer_font = ctk.CTkFont(family=BODY_FONT_FAMILY, size=BODY_FONT_SIZE_SMALL)
        self.button_font = ctk.CTkFont(family=BODY_FONT_FAMILY, size=BODY_FONT_SIZE_SMALL, weight='bold')
        self.metadata_label_font = ctk.CTkFont(family=BODY_FONT_FAMILY, size=BODY_FONT_SIZE_SMALL, weight='bold')
        self.metadata_entry_font = ctk.CTkFont(family=BODY_FONT_FAMILY, size=BODY_FONT_SIZE_SMALL)
        self.tab_font = ctk.CTkFont(family=BODY_FONT_FAMILY, size=BODY_FONT_SIZE_SMALL + 1, weight='bold')
        self.status_font = ctk.CTkFont(family=BODY_FONT_FAMILY, size=STATUS_FONT_SIZE)
        # self.link_entry_font removed

        # --- Initialize UI ---
        self._initialize_checklist_vars() # Use the reverted simple init
        self._create_menu()
        self._create_widgets()
        self.after(150, self._initial_checklist_build)

        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.update_title()
        self.status_var.set("Ready")

        if not OPENPYXL_AVAILABLE: messagebox.showwarning("Missing Library", "Excel export disabled.\npip install openpyxl")
        if not REPORTLAB_AVAILABLE: messagebox.showwarning("Missing Library", "PDF export disabled.\npip install reportlab")

    # --- Methods (Menu, Widgets, Initial Build, Utilities) ---
    def _initial_checklist_build(self):
        """Ensures checklist is built after window geometry is stable."""
        try:
            self.update_idletasks()
            if hasattr(self, 'checklist_frame'):
                self.checklist_frame.rebuild_checklist_ui()
        except Exception as e:
            print(f"Error during initial checklist build: {e}")
            messagebox.showerror("UI Error", "Critical error: Could not build the checklist view.")

    def _create_menu(self):
        """Creates the top menu bar (File, Help)."""
        menubar = tk.Menu(self)
        self.configure(menu=menubar)
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="New Checklist", command=self.new_checklist, accelerator="Ctrl+N")
        file_menu.add_command(label="Open Project (.json)...", command=self.load_project, accelerator="Ctrl+O")
        file_menu.add_command(label="Save Project", command=self.save_project, accelerator="Ctrl+S")
        file_menu.add_command(label="Save Project As... (.json)", command=self.save_project_as, accelerator="Ctrl+Shift+S")
        file_menu.add_separator()
        export_menu = tk.Menu(file_menu, tearoff=0)
        file_menu.add_cascade(label="Export Report As", menu=export_menu)
        ex_state_excel = tk.NORMAL if OPENPYXL_AVAILABLE else tk.DISABLED
        ex_state_pdf = tk.NORMAL if REPORTLAB_AVAILABLE else tk.DISABLED
        export_menu.add_command(label="Excel (.xlsx)...", command=lambda: self.export_data('excel'), state=ex_state_excel)
        export_menu.add_command(label="PDF (.pdf)...", command=lambda: self.export_data('pdf'), state=ex_state_pdf)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.on_closing)
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self.show_about)
        self.bind_all("<Control-n>", lambda event: self.new_checklist())
        self.bind_all("<Control-o>", lambda event: self.load_project())
        self.bind_all("<Control-s>", lambda event: self.save_project())
        self.bind_all("<Control-Shift-s>", lambda event: self.save_project_as())

    def _create_widgets(self):
        """Creates and grids all the main widgets in the window."""
        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)
        header_label = ctk.CTkLabel(self, text="Warehouse Safety Compliance Checklist", font=self.header_font, text_color=SECONDARY_COLOR, anchor="center")
        header_label.grid(row=0, column=0, pady=(15, 20), padx=20, sticky="ew")
        metadata_frame = ctk.CTkFrame(self, corner_radius=6, border_width=1, border_color=SECONDARY_COLOR, fg_color=BACKGROUND_COLOR)
        metadata_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 20))
        metadata_frame.columnconfigure((1, 4), weight=1, minsize=180)
        metadata_frame.columnconfigure(2, weight=0, minsize=40)
        fields = list(self.metadata_vars.keys()); row_num = 0; col_num = 0
        for i, field in enumerate(fields):
            label = ctk.CTkLabel(metadata_frame, text=f"{field}:", font=self.metadata_label_font, anchor="w", text_color=SECONDARY_COLOR)
            label.grid(row=row_num, column=col_num, sticky="w", padx=(15, 5), pady=7)
            if field == "Uploaded By Role":
                widget = ctk.CTkComboBox(metadata_frame, variable=self.metadata_vars[field], values=["Safety Champion", "Manager", "Other"], state="readonly", font=self.metadata_entry_font, width=150, button_color=PRIMARY_COLOR, border_width=1, border_color=PRIMARY_COLOR, dropdown_font=self.metadata_entry_font, dropdown_fg_color=BACKGROUND_COLOR, dropdown_hover_color=ACCENT_COLOR)
            else:
                widget = ctk.CTkEntry(metadata_frame, textvariable=self.metadata_vars[field], font=self.metadata_entry_font, width=150, border_width=1)
            widget.grid(row=row_num, column=col_num + 1, sticky="ew", padx=(0, 15), pady=7)
            if col_num == 0: col_num = 3
            else: col_num = 0; row_num += 1
        self.tabview = ctk.CTkTabview(self, corner_radius=6, border_width=1, border_color=PRIMARY_COLOR, segmented_button_selected_color=PRIMARY_COLOR, segmented_button_selected_hover_color=ACCENT_COLOR, segmented_button_unselected_color=BACKGROUND_COLOR, segmented_button_unselected_hover_color="#E0E0E0", text_color_disabled="gray60", fg_color=BACKGROUND_COLOR)
        self.tabview._segmented_button.configure(font=self.tab_font, text_color=SECONDARY_COLOR, selected_text_color=TEXT_ON_PRIMARY) # Corrected config
        self.tabview.grid(row=2, column=0, sticky="nsew", pady=(0,10), padx=20)
        tab_names = ["Checklist Items", "Near Miss Report", "Action Points", "General Links"]
        for name in tab_names:
            self.tabview.add(name); tab_frame = self.tabview.tab(name); tab_frame.configure(fg_color=BACKGROUND_COLOR); tab_frame.grid_rowconfigure(0, weight=1); tab_frame.grid_columnconfigure(0, weight=1)
        self.checklist_frame = ChecklistFrame(self.tabview.tab("Checklist Items"), self, self.checklist_data_vars) # Uses REVERTED ChecklistFrame
        self.checklist_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        self.near_miss_frame = NearMissFrame(self.tabview.tab("Near Miss Report"), self, self.near_miss_vars, self.near_miss_attachments)
        self.near_miss_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        self.action_points_frame = ActionPointsFrame(self.tabview.tab("Action Points"), self, self.action_points_text_var, "Action Points / Further Recommendations:")
        self.action_points_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        self.attachment_frame = LinkAttachmentFrame(self.tabview.tab("General Links"), self, self.general_attachments, "Overall Supporting Links")
        self.attachment_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        status_bar_frame = ctk.CTkFrame(self, height=28, corner_radius=0, fg_color="#EAEAEA", border_width=0); status_bar_frame.grid(row=3, column=0, sticky="ew", padx=0, pady=(10,0)); status_label = ctk.CTkLabel(status_bar_frame, textvariable=self.status_var, font=self.status_font, anchor="w", padx=15, text_color=TEXT_COLOR_LIGHT); status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

    # --- Data Handling Methods ---
    def _initialize_checklist_vars(self):
        """Creates/resets checklist data structure (answer ONLY)."""
        self.checklist_data_vars.clear()
        for _, questions in CHECKLIST_STRUCTURE:
            for qt, at, _ in questions:
                 if at == "yes_no": self.checklist_data_vars[qt] = tk.StringVar(value="")
                 elif at == "text": self.checklist_data_vars[qt] = tk.StringVar()

    def _clear_all_fields(self):
        """Clears all input fields and data structures."""
        try:
            for k, var in self.metadata_vars.items(): var.set("")
            self.metadata_vars["Report Date"].set(datetime.now().strftime('%Y-%m-%d'))
            self.metadata_vars["Report Month"].set(datetime.now().strftime('%B %Y'))
            self._initialize_checklist_vars()
            if hasattr(self, 'checklist_frame'): self.checklist_frame.rebuild_checklist_ui()
            for var in self.near_miss_vars.values(): var.set("")
            self.near_miss_attachments.clear()
            if hasattr(self, 'near_miss_frame'): self.near_miss_frame.update_attachment_list()
            self.action_points_text_var.set("")
            if hasattr(self, 'action_points_frame'): self.action_points_frame.clear_text()
            self.general_attachments.clear()
            if hasattr(self, 'attachment_frame'): self.attachment_frame.update_link_list()
        except Exception as e: print(f"Error during field clearing: {e}"); messagebox.showerror("Error", "Could not fully clear all fields.")

    def get_all_data(self):
        """Collects all data into a dictionary (simple checklist answers)."""
        action_points = self.action_points_frame.get_text() if hasattr(self, 'action_points_frame') else ""
        checklist_export_data = {k: v.get() for k, v in self.checklist_data_vars.items()}
        data = {"metadata": {k: v.get() for k, v in self.metadata_vars.items()},"checklist": checklist_export_data,"near_miss": {"details": {k: v.get() for k, v in self.near_miss_vars.items()},"attachments": self.near_miss_attachments},"action_points": action_points,"general_attachments": self.general_attachments}
        return data

    def load_data(self, data):
        """Populates UI elements from loaded data (simple checklist answers)."""
        try:
            self.status_var.set("Loading data...")
            loaded_meta = data.get("metadata", {}); [(self.metadata_vars[k].set(v)) for k,v in loaded_meta.items() if k in self.metadata_vars]
            loaded_checklist = data.get("checklist", {}) # Simple answers dict
            for q, a in loaded_checklist.items():
                if q in self.checklist_data_vars: self.checklist_data_vars[q].set(a)
            loaded_near_miss = data.get("near_miss", {}); nm_details = loaded_near_miss.get("details", {}); [(self.near_miss_vars[k].set(v)) for k,v in nm_details.items() if k in self.near_miss_vars]; self.near_miss_attachments.clear(); self.near_miss_attachments.extend(loaded_near_miss.get("attachments", [])); [getattr(self, nf,'update_attachment_list', lambda: None)() for nf in ['near_miss_frame'] if hasattr(self, nf)]
            action_points_text = data.get("action_points", ""); self.action_points_text_var.set(action_points_text); [getattr(self, af,'set_text', lambda t: None)(action_points_text) for af in ['action_points_frame'] if hasattr(self, af)]
            self.general_attachments.clear(); self.general_attachments.extend(data.get("general_attachments", [])); [getattr(self, lf,'update_link_list', lambda: None)() for lf in ['attachment_frame'] if hasattr(self, lf)]
            self.status_var.set("Data loaded successfully.")
        except Exception as e: messagebox.showerror("Load Error", f"Failed loading data: {e}"); self.status_var.set("Error loading.")

    # --- File Operations ---
    def new_checklist(self):
        current_data=self.get_all_data(); has_data=any(v for k,v in current_data['metadata'].items() if k not in ["Report Date","Report Month"]) or any(current_data['checklist'].values()) or any(current_data['near_miss']['details'].values()) or current_data['near_miss']['attachments'] or current_data['action_points'] or current_data['general_attachments']
        if self.project_file_path or has_data:
             if not messagebox.askyesno("Confirm New", "Discard current unsaved changes?", icon='warning'): return
        self.status_var.set("Creating new..."); self._clear_all_fields(); self.project_file_path=None; self.update_title(); self.status_var.set("New checklist ready."); [self.tabview.set("Checklist Items") if hasattr(self,'tabview') else None]

    def save_project(self):
        if not self.project_file_path: self.save_project_as()
        elif self.project_file_path:
            try: self._write_project_file(self.project_file_path)
            except Exception: self.status_var.set("Save failed.")

    # --- CORRECTED save_project_as (from previous fix) ---
    def save_project_as(self):
        """Prompts user for filename and saves the project."""
        file_path = None # Initialize
        try:
            wh_name = self.metadata_vars["Warehouse Name"].get().replace(" ", "_") or "UnknownWH"
            rep_date = self.metadata_vars["Report Date"].get() or datetime.now().strftime('%Y%m%d')
            initial_file = f"SafetyChecklist_{wh_name}_{rep_date}.json"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("Checklist Project Files", "*.json"), ("All Files", "*.*")],
                title="Save Project As",
                initialfile=initial_file,
                initialdir=os.path.dirname(self.project_file_path) if self.project_file_path else os.getcwd()
            )
            # Check if user cancelled
            if not file_path:
                self.status_var.set("Save cancelled.")
                return # Exit the function if cancelled

            # ---- Code to execute if file_path IS valid ----
            self.project_file_path = file_path # Set the path *before* writing
            self._write_project_file(file_path) # Attempt to write the file (this might fail)
            self.update_title() # Update title only if write succeeds

        except Exception as e: # Catch errors from _write_project_file or file dialog
             messagebox.showerror("Save Error", f"Failed saving project: {e}")
             self.status_var.set("Save error.")
             # Reset path only if Save As failed to get a path initially,
             # or if _write_project_file failed (it resets path itself)
             if not self.project_file_path or self.project_file_path != file_path:
                 self.project_file_path = None
             self.update_title()
    # --- End CORRECTED save_project_as ---

    def _write_project_file(self, file_path):
        self.status_var.set(f"Saving: {os.path.basename(file_path)}..."); data_to_save=self.get_all_data()
        try:
            with open(file_path,'w',encoding='utf-8') as f: json.dump(data_to_save, f, indent=4, ensure_ascii=False)
            self.status_var.set(f"Saved: {os.path.basename(file_path)}")
        except IOError as e: messagebox.showerror("File Write Error",f"Could not write file:\n{file_path}\nError: {e}"); self.status_var.set("Error saving."); self.project_file_path = None; self.update_title(); raise
        except Exception as e: messagebox.showerror("Save Error",f"Unexpected save error:\n{e}"); self.status_var.set("Error saving."); self.project_file_path = None; self.update_title(); raise

    # --- CORRECTED load_project Function ---
    def load_project(self):
        """Loads project data from a JSON file."""
        file_path = None
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("Checklist Project Files", "*.json"), ("All Files", "*.*")],
                title="Open Project",
                initialdir=os.path.dirname(self.project_file_path) if self.project_file_path else os.getcwd()
            )
            if not file_path:
                self.status_var.set("Open cancelled.")
                return

            self.status_var.set(f"Loading: {os.path.basename(file_path)}...")
            with open(file_path, 'r', encoding='utf-8') as f:
                loaded_data = json.load(f)

            self._clear_all_fields() # Clear before loading
            self.load_data(loaded_data) # Load into UI

            self.project_file_path = file_path # Update path
            self.update_title()
            if hasattr(self, 'tabview'): self.tabview.set("Checklist Items")
            self.status_var.set(f"Loaded: {os.path.basename(file_path)}")

        except FileNotFoundError: messagebox.showerror("Load Error", f"File not found:\n{file_path or '?'}"); self.status_var.set("Error: File not found.")
        except json.JSONDecodeError: messagebox.showerror("Load Error", f"Invalid project file format:\n{file_path or '?'}"); self.status_var.set("Error: Invalid project file.")
        except Exception as e: messagebox.showerror("Load Error", f"Error loading project:\n{e}"); self.status_var.set("Load error."); self.project_file_path = None; self.update_title()
    # --- End CORRECTED load_project Function ---

    def validate_for_export(self):
        required_meta=["Warehouse Name","Location","Uploaded By Name","Uploaded By Role"]; missing=[f for f in required_meta if not self.metadata_vars[f].get()];
        if missing: messagebox.showerror("Missing Info","Fill required fields:\n- "+"\n- ".join(missing)); return False
        return True

    def export_data(self, format_type):
        if not self.validate_for_export(): return
        data=self.get_all_data(); wh_name=self.metadata_vars["Warehouse Name"].get().replace(" ","_") or "UnknownWH"; rep_date=self.metadata_vars["Report Date"].get() or datetime.now().strftime('%Y%m%d'); default_filename=f"SafetyReport_{wh_name}_{rep_date}"; export_successful=False; file_path=None
        try:
            if format_type=='excel':
                if not OPENPYXL_AVAILABLE: messagebox.showerror("Missing Library","Excel export needs 'openpyxl'."); return
                file_path=filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=[("Excel Files","*.xlsx")],initialfile=default_filename + ".xlsx",title="Export Report as Excel")
                if file_path: self.status_var.set("Exporting Excel..."); self._export_to_excel(data, file_path); export_successful=True
            elif format_type=='pdf':
                 if not REPORTLAB_AVAILABLE: messagebox.showerror("Missing Library","PDF export needs 'reportlab'."); return
                 file_path=filedialog.asksaveasfilename(defaultextension=".pdf",filetypes=[("PDF Files","*.pdf")],initialfile=default_filename + ".pdf",title="Export Report as PDF")
                 if file_path: self.status_var.set("Exporting PDF..."); self._export_to_pdf(data, file_path); export_successful=True

            if export_successful and file_path: messagebox.showinfo("Export Successful",f"Report exported:\n{file_path}\n\nIMPORTANT: Ensure links accessible & send report to admin."); self.status_var.set(f"Exported: {os.path.basename(file_path)}")
            elif file_path: self.status_var.set("Export failed.") # Error handled in helper
            else: self.status_var.set("Export cancelled.")
        except Exception as e: messagebox.showerror("Export Error",f"Export error ({format_type.upper()}):\n{e}"); self.status_var.set(f"Error exporting.")

    # --- Export Helper Methods ---
    def _export_to_excel(self, data, file_path):
        """Exports data to Excel (simple checklist format - Corrected)."""
        try:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Safety Checklist Report"
            # --- Styling ---
            H_FONT=OpenpyxlFont(name='Arial Black',size=16,bold=True,color="FF14467C"); MH_FONT=OpenpyxlFont(name='Arial',size=14,bold=True,color="FF14467C"); S_FONT=OpenpyxlFont(name='Arial',size=12,bold=True,color="FF39B54A"); Q_FONT=OpenpyxlFont(name='Arial',size=11,bold=True); A_FONT=OpenpyxlFont(name='Arial',size=11); LINK_FONT=OpenpyxlFont(name='Arial',size=10,italic=True,underline='single',color='FF0000FF'); ML_FONT=OpenpyxlFont(name='Arial',size=11,bold=True); MV_FONT=OpenpyxlFont(name='Arial',size=11); WRAP_ALIGN=Alignment(wrap_text=True,vertical='top',horizontal='left'); CENTER_ALIGN=Alignment(vertical='center',horizontal='center'); BORDER_SIDE=Side(border_style="thin",color="FFDDDDDD"); BORDER=Border(left=BORDER_SIDE,right=BORDER_SIDE,top=BORDER_SIDE,bottom=BORDER_SIDE); FILL=PatternFill(start_color="FFEAEAEA",end_color="FFEAEAEA",fill_type="solid")
            ws.column_dimensions['A'].width = 50; ws.column_dimensions['B'].width = 70; row = 1
            # --- Header ---
            c = ws.cell(row, 1, "Warehouse Safety Compliance Report"); c.font = H_FONT; ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2); c.alignment = CENTER_ALIGN; row += 2
            # --- Metadata ---
            c = ws.cell(row, 1, "Report Information"); c.font = MH_FONT; ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2); c.fill = FILL; row += 1
            for k, v in data['metadata'].items(): ca = ws.cell(row, 1, f"{k}:"); ca.font = ML_FONT; ca.border = BORDER; cb = ws.cell(row, 2, v); cb.font = MV_FONT; cb.alignment = WRAP_ALIGN; cb.border = BORDER; row += 1
            row += 1
            # --- Checklist Items ---
            c = ws.cell(row, 1, "Checklist Items"); c.font = MH_FONT; ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2); c.fill = FILL; row += 1
            ws.cell(row, 1, "Question").font = Q_FONT; ws.cell(row, 1).border = BORDER; ws.cell(row, 2, "Answer").font = A_FONT; ws.cell(row, 2).border = BORDER; row += 1
            for section_title, questions in CHECKLIST_STRUCTURE:
                 cs = ws.cell(row, 1, section_title); cs.font = S_FONT; ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2); row += 1
                 for qt, _, m in questions:
                     qd = f"{qt}{' *' if m else ''}"; a = data['checklist'].get(qt, "[N/A]")
                     ca = ws.cell(row, 1, qd); ca.font = Q_FONT; ca.alignment = WRAP_ALIGN; ca.border = BORDER
                     cb = ws.cell(row, 2, a if a else "[N/A]"); cb.font = A_FONT; cb.alignment = WRAP_ALIGN; cb.border = BORDER
                     row += 1
            row += 1 # Spacer
            # --- Near Miss Report ---
            c=ws.cell(row,1,"Near Miss Report"); c.font=MH_FONT; ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2); c.fill=FILL; row+=1
            nm_details=data['near_miss']['details']
            if any(nm_details.values()):
                for k, v in nm_details.items(): ca=ws.cell(row,1,f"{k.replace('_',' ')}:"); ca.font=ML_FONT; ca.border=BORDER; cb=ws.cell(row,2,v if v else "[N/A]"); cb.font=A_FONT; cb.alignment=WRAP_ALIGN; cb.border=BORDER; row+=1
                nm_att=data['near_miss']['attachments']; cal=ws.cell(row,1,"Near Miss Evidence Links:"); cal.font=ML_FONT; cal.border=BORDER
                if nm_att:
                     if len(nm_att) > 1: ws.merge_cells(start_row=row,start_column=1,end_row=row+len(nm_att)-1,end_column=1) # Merge label only if multiple links
                     else: ws.cell(row=row, column=1).border = BORDER # Ensure border on single label cell
                     start_r=row
                     # --- Corrected Loop ---
                     for i,url in enumerate(nm_att):
                         current_row_for_link = start_r + i
                         cell = ws.cell(row=current_row_for_link, column=2, value=url)
                         cell.font = LINK_FONT
                         cell.border = BORDER
                         if url and url.startswith("http"): cell.hyperlink = url
                     # --- End Corrected Loop ---
                     row+=len(nm_att) # Increment row counter AFTER loop
                else: can=ws.cell(row,2,"[None]"); can.font=LINK_FONT; can.border=BORDER; row+=1
            else: cnn=ws.cell(row,1,"[No Near Miss Recorded]"); cnn.font=A_FONT; ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2); cnn.border=BORDER; row+=1
            row+=1 # Spacer
            # --- Action Points ---
            c=ws.cell(row,1,"Action Points / Recommendations"); c.font=MH_FONT; ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2); c.fill=FILL; row+=1
            ap_text=data['action_points']; cap=ws.cell(row,1,ap_text if ap_text else "[None]"); cap.font=A_FONT; cap.alignment=WRAP_ALIGN; cap.border=BORDER; ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2); row+=2
            # --- General Evidence Links ---
            c=ws.cell(row,1,"Overall Supporting Links"); c.font=MH_FONT; ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2); c.fill=FILL; row+=1
            gen_att=data['general_attachments']
            if gen_att:
                 # --- Corrected Loop ---
                 for url in gen_att:
                     cell=ws.cell(row,1,url)
                     cell.font = LINK_FONT
                     cell.border = BORDER
                     if url and url.startswith("http"): cell.hyperlink = url
                     ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2)
                     row+=1
                 # --- End Corrected Loop ---
            else: cga=ws.cell(row,1,"[None]"); cga.font=LINK_FONT; cga.border=BORDER; ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=2); row+=1
            # --- Save ---
            wb.save(file_path)
        except PermissionError: messagebox.showerror("Save Error",f"Permission denied:\n'{os.path.basename(file_path)}'\nIs it open?"); raise
        except Exception as e: messagebox.showerror("Excel Export Error",f"Error saving Excel:\n{e}"); raise

    def _export_to_pdf(self, data, file_path):
        """Exports data to PDF (simple checklist format)."""
        try:
            doc=SimpleDocTemplate(file_path,pagesize=(8.5*inch,11*inch),leftMargin=0.6*inch,rightMargin=0.6*inch,topMargin=0.6*inch,bottomMargin=0.6*inch); styles=getSampleStyleSheet(); story=[]
            # Styles...
            try: from reportlab.pdfbase.ttfonts import TTFont; from reportlab.pdfbase import pdfmetrics; header_font_name='Arial-Black'
            except: header_font_name='Helvetica-Bold'
            styles.add(ParagraphStyle(name='MainHeader',fontName=header_font_name,fontSize=18,textColor=colors.HexColor(SECONDARY_COLOR),alignment=TA_CENTER,spaceAfter=10)); styles.add(ParagraphStyle(name='SubHeader',parent=styles['Normal'],alignment=TA_CENTER,fontSize=10,textColor=colors.dimgrey,spaceAfter=15)); styles.add(ParagraphStyle(name='MetaHeader',fontName='Helvetica-Bold',fontSize=14,textColor=colors.HexColor(SECONDARY_COLOR),spaceBefore=12,spaceAfter=6,keepWithNext=1)); styles.add(ParagraphStyle(name='MetaLabel',fontName='Helvetica-Bold',fontSize=10,textColor=colors.black)); styles.add(ParagraphStyle(name='MetaValue',parent=styles['Normal'],fontSize=10,leftIndent=15,spaceAfter=3)); styles.add(ParagraphStyle(name='SectionHeaderPDF',fontName='Helvetica-Bold',fontSize=12,textColor=colors.HexColor(PRIMARY_COLOR),spaceBefore=15,spaceAfter=8,keepWithNext=1,backgroundColor=colors.HexColor("#F0F0F0"),padding=4,borderRadius=3)); styles.add(ParagraphStyle(name='QuestionStylePDF',parent=styles['Normal'],fontName='Helvetica-Bold',fontSize=10,textColor=colors.black,spaceBefore=6,leftIndent=10,allowWidows=1,allowOrphans=1,keepWithNext=0)); styles.add(ParagraphStyle(name='AnswerStylePDF',parent=styles['Normal'],fontName='Helvetica',fontSize=10,textColor=colors.darkslategray,leftIndent=25,spaceAfter=5,wordWrap='CJK',leading=12, keepWithNext=1)); styles.add(ParagraphStyle(name='AnswerStyleEmptyPDF',parent=styles['AnswerStylePDF'],textColor=colors.HexColor(DARK_GREY),fontName='Helvetica-Oblique')); styles.add(ParagraphStyle(name='NMFieldLabelPDF',parent=styles['MetaLabel'],leftIndent=10)); styles.add(ParagraphStyle(name='NMFieldValuePDF',parent=styles['MetaValue'],leftIndent=25)); styles.add(ParagraphStyle(name='AttachLabelPDF',parent=styles['MetaLabel'],leftIndent=10,spaceBefore=5,keepWithNext=1)); styles.add(ParagraphStyle(name='AttachmentLinkPDF',parent=styles['Normal'],fontSize=9,leftIndent=25,textColor=colors.blue,spaceAfter=2))
            def pdf_escape(text): return text.replace('&','&').replace('<','<').replace('>','>').replace('\n','<br/>') if text else ""
            def create_link_paragraph(url, style=styles['AttachmentLinkPDF']):
                 if url and url.startswith("http"): escaped_url=pdf_escape(url); display_url=escaped_url if len(escaped_url) < 70 else escaped_url[:67]+"..."; return Paragraph(f'<link href="{escaped_url}">{display_url}</link>', style)
                 else: return Paragraph(pdf_escape(url) if url else "[Invalid Link]", styles['AnswerStyleEmptyPDF'])
            # Story build...
            story.append(Paragraph("Warehouse Safety Compliance Report",styles['MainHeader'])); story.append(Paragraph(f"Date: {pdf_escape(data['metadata'].get('Report Date','N/A'))} | WH: {pdf_escape(data['metadata'].get('Warehouse Name','N/A'))} | Loc: {pdf_escape(data['metadata'].get('Location','N/A'))}",styles['SubHeader'])); story.append(Paragraph("Report Information",styles['MetaHeader']))
            meta_data_table=[]; meta=data['metadata']; fields_ordered=["Warehouse Name","Location","Report Date","Report Month","Uploaded By Name","Uploaded By Role","Uploaded By Emp ID","Uploaded By Email","Manager Name"]
            for i in range(0,len(fields_ordered),2): key1=fields_ordered[i]; val1=pdf_escape(meta.get(key1,'')); p1_label=Paragraph(f"<b>{key1}:</b>",styles['MetaLabel']); p1_value=Paragraph(val1 if val1 else "[N/A]",styles['MetaValue']); p2_label,p2_value=("","") if i+1>=len(fields_ordered) else (key2:=fields_ordered[i+1],val2:=pdf_escape(meta.get(key2,'')),Paragraph(f"<b>{key2}:</b>",styles['MetaLabel']),Paragraph(val2 if val2 else "[N/A]",styles['MetaValue'])); meta_data_table.append([p1_label,p1_value,p2_label,p2_value])
            table=Table(meta_data_table,colWidths=[1.5*inch,2.2*inch,1.5*inch,2.2*inch]); table.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),2)])); story.append(table); story.append(Spacer(1,0.2*inch))
            story.append(Paragraph("Checklist Items",styles['MetaHeader']))
            for section_title, questions in CHECKLIST_STRUCTURE:
                 section_items=[Paragraph(section_title,styles['SectionHeaderPDF'])]
                 for qt, _, m in questions:
                     qd=f"{qt}{' *' if m else ''}"; a=data['checklist'].get(qt)
                     p_q=Paragraph(pdf_escape(qd),styles['QuestionStylePDF'])
                     p_a=Paragraph(pdf_escape(a),styles['AnswerStylePDF']) if a else Paragraph("[N/A]",styles['AnswerStyleEmptyPDF'])
                     section_items.extend([p_q, p_a])
                 story.append(KeepTogether(section_items))
            story.append(PageBreak()); story.append(Paragraph("Near Miss Report",styles['MetaHeader'])); nm_details=data['near_miss']['details']
            if any(nm_details.values()):
                 nm_section_content=[]; field_map={"Incident Date":"Date","Incident Location":"Location","Description":"Description","Immediate Action":"Action","Prevention Suggestion":"Prevention"}
                 for k,lbl in field_map.items(): v=pdf_escape(nm_details.get(k,'')); p_l=Paragraph(f"<b>{lbl}:</b>",styles['NMFieldLabelPDF']); p_v=Paragraph(v if v else "[N/A]",styles['NMFieldValuePDF']); nm_section_content.extend([p_l,p_v])
                 nm_att=data['near_miss']['attachments']; nm_section_content.append(Paragraph("<b>Evidence Links (Near Miss):</b>",styles['AttachLabelPDF']))
                 nm_section_content.extend([create_link_paragraph(url) for url in nm_att]) if nm_att else nm_section_content.append(Paragraph("[None]",styles['AnswerStyleEmptyPDF']))
                 story.append(KeepTogether(nm_section_content))
            else: story.append(Paragraph("[No Near Miss Recorded]",styles['AnswerStyleEmptyPDF']))
            story.append(Spacer(1,0.2*inch)); story.append(Paragraph("Action Points / Recommendations",styles['MetaHeader'])); ap_text=data['action_points']; story.append(Paragraph(pdf_escape(ap_text),styles['AnswerStylePDF']) if ap_text else Paragraph("[None]",styles['AnswerStyleEmptyPDF'])); story.append(Spacer(1,0.2*inch)); story.append(Paragraph("Overall Supporting Links",styles['MetaHeader'])); gen_att=data['general_attachments']; story.extend([create_link_paragraph(url) for url in gen_att]) if gen_att else story.append(Paragraph("[None]",styles['AnswerStyleEmptyPDF']))
            # Build
            doc.build(story)
        except PermissionError: messagebox.showerror("Save Error",f"Permission denied:\n'{os.path.basename(file_path)}'\nIs it open?"); raise
        except Exception as e: messagebox.showerror("PDF Export Error",f"Failed building PDF:\n{e}"); raise

    # --- Utility Methods ---
    def update_title(self): base_title="Warehouse Safety Checklist"; proj_name=os.path.basename(self.project_file_path) if self.project_file_path else "New Project"; self.title(f"{base_title} - {proj_name}")
    def show_about(self): messagebox.showinfo("About Warehouse Safety Checklist","Warehouse Safety Compliance App (v8.2)\n\nVersion: 8.2\nUI Refined. Checklist links removed.\n\nEnsure shared links accessible.")
    def on_closing(self):
        if messagebox.askyesno("Exit Application","Are you sure you want to exit?",icon='question'): self.destroy()


# ==============================================================================
# Frame Classes (Using CustomTkinter Widgets - Reverted ChecklistFrame)
# ==============================================================================

# --- Checklist Frame (Reverted to 2 columns) ---
class ChecklistFrame(ctk.CTkScrollableFrame):
    """Scrollable frame for the main checklist questions and answers."""
    def __init__(self, master, app_controller, checklist_data_vars):
        super().__init__(master, corner_radius=5, fg_color=BACKGROUND_COLOR)
        self.app = app_controller
        self.checklist_data_vars = checklist_data_vars # Should be simple {q: var} dict
        self.question_widgets = {}
        self._scrollbar.configure(width=16, button_color=PRIMARY_COLOR, button_hover_color=ACCENT_COLOR)
        # Configure internal grid columns (BACK TO 2 COLUMNS)
        self.grid_columnconfigure(0, weight=3, uniform="checklist_cols") # Question column
        self.grid_columnconfigure(1, weight=2, uniform="checklist_cols") # Answer column

    def rebuild_checklist_ui(self):
        """Clears and rebuilds the checklist UI (simple Q & A)."""
        # Clear previous widgets reliably
        if hasattr(self, '_canvas'):
            for widget in list(self._canvas.winfo_children()): # Iterate over a copy
                if widget.master == self._canvas:
                    try: widget.destroy()
                    except tk.TclError: pass # Ignore if already gone
        self.question_widgets.clear()

        current_row = 0
        wrap_len = 450 # Adjust wrap length as needed

        for section_index, (section_title, questions) in enumerate(CHECKLIST_STRUCTURE):
            try:
                section_label = ctk.CTkLabel(self, text=section_title, font=self.app.section_header_font, anchor="w", text_color=SECONDARY_COLOR)
                section_label.grid(row=current_row, column=0, columnspan=2, sticky="ew", pady=(18 if section_index > 0 else 5, 6), padx=10) # Spans 2 cols
                current_row += 1
                sep = ctk.CTkFrame(self, height=2, fg_color=PRIMARY_COLOR)
                sep.grid(row=current_row, column=0, columnspan=2, sticky='ew', padx=10, pady=(0, 10)) # Spans 2 cols
                current_row += 1

                for question_text, answer_type, mandatory in questions:
                    if question_text not in self.checklist_data_vars:
                        print(f"ERROR: Variable missing for '{question_text}'. Skipping.")
                        continue

                    answer_var = self.checklist_data_vars[question_text]
                    answer_widget = None; question_label = None

                    # Column 0: Question Label
                    try: q_display_text = f"{question_text}{' *' if mandatory else ''}"; question_label = ctk.CTkLabel(self, text=q_display_text, font=self.app.question_font, anchor="nw", wraplength=wrap_len, justify="left"); question_label.grid(row=current_row, column=0, sticky="nw", padx=(15, 10), pady=5)
                    except Exception as label_e: print(f"ERROR creating label for '{question_text}': {label_e}"); continue

                    # Column 1: Answer Widget
                    try:
                        if answer_type == "yes_no":
                            radio_frame = ctk.CTkFrame(self, fg_color="transparent"); args = {"master":radio_frame, "variable":answer_var, "font":self.app.answer_font,"radiobutton_width":18, "radiobutton_height":18,"fg_color":PRIMARY_COLOR,"hover_color":ACCENT_COLOR,"border_color":SECONDARY_COLOR}
                            rb_yes = ctk.CTkRadioButton(text="Yes", value="Yes", **args); rb_yes.pack(side=tk.LEFT, padx=(0, 20)); rb_no = ctk.CTkRadioButton(text="No", value="No", **args); rb_no.pack(side=tk.LEFT, padx=(0, 20)); rb_na = ctk.CTkRadioButton(text="N/A", value="N/A", **args); rb_na.pack(side=tk.LEFT, padx=(0, 15)); answer_widget = radio_frame
                        elif answer_type == "text": entry = ctk.CTkEntry(self, textvariable=answer_var, font=self.app.answer_font, width=280, border_width=1, corner_radius=5); answer_widget = entry
                        if answer_widget: answer_widget.grid(row=current_row, column=1, sticky="ew", padx=10, pady=5)
                    except Exception as ans_e: print(f"ERROR creating answer widget for '{question_text}': {ans_e}"); continue

                    self.question_widgets[question_text] = (question_label, answer_widget)
                    current_row += 1
            except Exception as section_e: print(f"ERROR processing section '{section_title}': {section_e}")

# --- Near Miss Frame (Unchanged) ---
class NearMissFrame(ctk.CTkFrame):
    def __init__(self, master, app_controller, near_miss_data_vars, near_miss_attachments_ref):
        super().__init__(master, fg_color=BACKGROUND_COLOR); self.app=app_controller; self.near_miss_vars=near_miss_data_vars; self.attachments_ref=near_miss_attachments_ref; self.grid_columnconfigure(1, weight=1)
        fields=list(self.near_miss_vars.keys()); field_labels={"Incident Date":"Date:","Incident Location":"Location:","Description":"Description:","Immediate Action":"Action Taken:","Prevention Suggestion":"Prevention:"}; row_num=0; self.detail_widgets={}
        for key in fields:
            label_text=field_labels.get(key, key + ":"); label=ctk.CTkLabel(self, text=label_text, font=self.app.question_font, anchor="nw", text_color=SECONDARY_COLOR); label.grid(row=row_num, column=0, sticky="nw", padx=15, pady=(12,2))
            if key in ["Description", "Immediate Action", "Prevention Suggestion"]: widget=ctk.CTkTextbox(self, wrap=tk.WORD, height=75, font=self.app.answer_font, border_width=1, corner_radius=5, border_color=PRIMARY_COLOR); widget.insert("1.0", self.near_miss_vars[key].get()); widget.bind("<FocusOut>", lambda ev, k=key, w=widget: self.near_miss_vars[k].set(w.get("1.0", "end-1c")), add="+"); self.near_miss_vars[key].trace_add("write", lambda n, i, m, k=key, w=widget: self._update_textbox_content(w, self.near_miss_vars[k]))
            else: widget=ctk.CTkEntry(self, textvariable=self.near_miss_vars[key], font=self.app.answer_font, width=200, border_width=1)
            widget.grid(row=row_num, column=1, sticky="ew", padx=15, pady=(10,2)); self.detail_widgets[key]=widget; row_num+=1
        ctk.CTkFrame(self, height=1, fg_color="gray80").grid(row=row_num, column=0, columnspan=2, sticky='ew', pady=25); row_num+=1
        ctk.CTkLabel(self, text="Evidence Links (Near Miss):", font=self.app.question_font, text_color=SECONDARY_COLOR).grid(row=row_num, column=0, columnspan=2, sticky="w", pady=(0, 8), padx=15); row_num+=1
        self.link_frame=LinkAttachmentSubFrame(self, self.app, self.attachments_ref, is_near_miss=True); self.link_frame.grid(row=row_num, column=0, columnspan=2, sticky="nsew", padx=15, pady=(0, 15)); self.rowconfigure(row_num, weight=1)
    def _update_textbox_content(self, textbox, string_var): ct=textbox.get("1.0","end-1c"); nt=string_var.get(); [(textbox.delete("1.0","end"), textbox.insert("1.0",nt)) if ct!=nt else None]
    def update_attachment_list(self): self.link_frame.update_link_list()

# --- Action Points Frame (Unchanged) ---
class ActionPointsFrame(ctk.CTkFrame):
    def __init__(self, master, app_controller, text_variable, label_text):
        super().__init__(master, fg_color=BACKGROUND_COLOR); self.app=app_controller; self.text_variable=text_variable; self.textbox=None
        self.grid_rowconfigure(1, weight=1); self.grid_columnconfigure(0, weight=1)
        label=ctk.CTkLabel(self, text=label_text, font=self.app.section_header_font, anchor="w", text_color=SECONDARY_COLOR); label.grid(row=0, column=0, sticky="ew", padx=15, pady=(15, 10))
        self.textbox=ctk.CTkTextbox(self, wrap=tk.WORD, font=self.app.answer_font, border_width=1, corner_radius=5, border_color=PRIMARY_COLOR); self.textbox.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0, 15))
        self.set_text(self.text_variable.get()); self.textbox.bind("<FocusOut>", self._update_variable, add="+"); self.text_variable.trace_add("write", self._on_var_write)
    def _update_variable(self, event=None):
        if self.textbox: current_text=self.textbox.get("1.0","end-1c"); [self.text_variable.set(current_text) if self.text_variable.get()!=current_text else None]
    def _on_var_write(self, *args):
        if self.textbox: current_text=self.textbox.get("1.0","end-1c"); new_text=self.text_variable.get()
            if current_text!=new_text:
                try: modified=self.textbox.edit_modified(); self.textbox.delete("1.0","end"); self.textbox.insert("1.0",new_text); self.textbox.edit_modified(False)
                except tk.TclError: pass
    def get_text(self): return self.textbox.get("1.0","end-1c") if self.textbox else ""
    def set_text(self, text): [(self.textbox.delete("1.0","end"), self.textbox.insert("1.0",text), self.textbox.edit_modified(False)) if self.textbox else None]
    def clear_text(self): [(self.textbox.delete("1.0","end"), self.textbox.edit_modified(False)) if self.textbox else None]

# --- Link Attachment Frame (Unchanged) ---
class LinkAttachmentFrame(ctk.CTkFrame):
     def __init__(self, master, app_controller, attachment_list_ref, label_text):
         super().__init__(master, fg_color=BACKGROUND_COLOR); self.app=app_controller; self.attachments_ref=attachment_list_ref; self.label_text=label_text
         self.grid_rowconfigure(1, weight=1); self.grid_columnconfigure(0, weight=1)
         ctk.CTkLabel(self, text=self.label_text, font=self.app.section_header_font, text_color=SECONDARY_COLOR).grid(row=0, column=0, sticky="w", pady=(15, 10), padx=15)
         self.link_subframe = LinkAttachmentSubFrame(self, self.app, self.attachments_ref, is_near_miss=False); self.link_subframe.grid(row=1, column=0, sticky="nsew", padx=15, pady=(0,15))
     def update_link_list(self): self.link_subframe.update_link_list()

# --- Link Attachment Sub-Frame (Unchanged) ---
class LinkAttachmentSubFrame(ctk.CTkFrame):
    def __init__(self, master, app_controller, attachment_list_ref, is_near_miss=False):
        super().__init__(master, fg_color="transparent"); self.app=app_controller; self.attachments_ref=attachment_list_ref; self.is_near_miss=is_near_miss
        self.grid_rowconfigure(0, weight=1); self.grid_columnconfigure(0, weight=1)
        list_container=ctk.CTkFrame(self, border_width=1, corner_radius=5, border_color=PRIMARY_COLOR); list_container.grid(row=0, column=0, sticky="nsew", pady=(0, 10)); list_container.grid_rowconfigure(0, weight=1); list_container.grid_columnconfigure(0, weight=1)
        self.link_list_frame=ctk.CTkScrollableFrame(list_container, fg_color="transparent", scrollbar_button_color=PRIMARY_COLOR, scrollbar_button_hover_color=ACCENT_COLOR); self.link_list_frame.grid(row=0, column=0, sticky="nsew"); self.link_list_frame.grid_columnconfigure(0, weight=1)
        button_frame=ctk.CTkFrame(self, fg_color="transparent"); button_frame.grid(row=1, column=0, sticky="w", pady=(0,5))
        add_button=ctk.CTkButton(button_frame, text="Add Link...", command=self.add_link, font=self.app.button_font, width=120, fg_color=PRIMARY_COLOR, hover_color=ACCENT_COLOR, text_color=TEXT_ON_PRIMARY); add_button.pack(side=tk.LEFT, padx=(0, 10))
        self.remove_button=ctk.CTkButton(button_frame, text="Remove Selected", command=self.remove_selected_link, font=self.app.button_font, width=120, state=tk.DISABLED, fg_color=SECONDARY_COLOR, hover_color="#2C5D8F", text_color=TEXT_ON_SECONDARY); self.remove_button.pack(side=tk.LEFT)
        self.selected_link_widget=None; self.link_widgets=[]
        self.update_link_list()
    def add_link(self): context="Near Miss" if self.is_near_miss else "General"; url=simpledialog.askstring(f"Add {context} Evidence Link","Paste URL (Ensure sharing permissions are correct!):",parent=self.app)
        if url:
             if not url.startswith(("http://","https://")): messagebox.showwarning("Invalid Link","URL must start with http:// or https://",parent=self.app); return
             if url not in self.attachments_ref: self.attachments_ref.append(url); self.update_link_list(); self.app.status_var.set(f"{context} link added.")
             else: messagebox.showinfo("Duplicate Link",f"Link already added for {context.lower()} evidence.",parent=self.app)
    def remove_selected_link(self): context="Near Miss" if self.is_near_miss else "General"
        if self.selected_link_widget and hasattr(self.selected_link_widget,"_url_reference"): url_to_remove=self.selected_link_widget._url_reference
            if url_to_remove in self.attachments_ref:
                 try: self.attachments_ref.remove(url_to_remove); self.update_link_list(); self.app.status_var.set(f"{context} link removed.")
                 except ValueError: messagebox.showerror("Error","Could not remove link (list mismatch).",parent=self.app); self.update_link_list()
            else: messagebox.showerror("Error","Selected link not found.",parent=self.app); self.update_link_list()
        else: messagebox.showwarning("No Selection","Please click on a link to select it for removal.",parent=self.app)
    def _on_link_select(self, clicked_widget, url):
        try: default_text_color=ctk.ThemeManager.theme["CTkLabel"]["text_color"]
        except KeyError: default_text_color = ("#000000", "#FFFFFF") # Fallback
        try: select_text_color=ctk.ThemeManager.theme["CTkButton"]["text_color"]
        except KeyError: select_text_color = ("#000000", "#FFFFFF") # Fallback

        if self.selected_link_widget and self.selected_link_widget!=clicked_widget:
            try: self.selected_link_widget.configure(fg_color="transparent",text_color=default_text_color)
            except tk.TclError: pass
        clicked_widget.configure(fg_color=ACCENT_COLOR,text_color=select_text_color); self.selected_link_widget=clicked_widget; self.remove_button.configure(state=tk.NORMAL)
    def update_link_list(self):
         for widget in list(self.link_list_frame.winfo_children()): # Use list() for safe iteration while destroying
             if isinstance(widget,ctk.CTkButton):
                 try: widget.destroy()
                 except: pass # Ignore if already destroyed
         self.link_widgets.clear(); self.selected_link_widget=None; self.remove_button.configure(state=tk.DISABLED)
         try: default_text_color = ctk.ThemeManager.theme["CTkLabel"]["text_color"]
         except KeyError: default_text_color = ("#000000", "#FFFFFF")
         for i, url in enumerate(self.attachments_ref): link_widget=ctk.CTkButton(self.link_list_frame,text=url,font=self.app.answer_font,anchor="w",fg_color="transparent",text_color=default_text_color,hover=False,corner_radius=3); link_widget._url_reference=url; link_widget.configure(command=lambda w=link_widget,u=url: self._on_link_select(w,u)); link_widget.grid(row=i,column=0,sticky="ew",padx=5,pady=1); self.link_widgets.append(link_widget)


# ==============================================================================
# Main Execution Block
# ==============================================================================
if __name__ == "__main__":
    try:
        app = WarehouseSafetyApp()
        app.mainloop()
    except Exception as e:
        print(f"FATAL ERROR: Could not start application.\n{e}")
        try: root = tk.Tk(); root.withdraw(); messagebox.showerror("Application Startup Error",f"Could not start:\n\n{e}"); root.destroy()
        except: pass # Ignore errors during fallback